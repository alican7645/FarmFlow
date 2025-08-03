import os
import sqlite3
import logging
from flask import Flask, render_template, request, redirect, url_for, flash, make_response, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from functools import wraps

# Import models
from models import db, User, LoginAttempt

# Configure logging
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "sera-yonetim-secret-key")

# PostgreSQL Database configuration
database_url = os.environ.get('DATABASE_URL')
if not database_url:
    # Fallback to individual PostgreSQL environment variables
    host = os.environ.get('PGHOST', 'localhost')
    port = os.environ.get('PGPORT', '5432')
    database = os.environ.get('PGDATABASE', 'postgres')
    user = os.environ.get('PGUSER', 'postgres')
    password = os.environ.get('PGPASSWORD', '')
    database_url = f'postgresql://{user}:{password}@{host}:{port}/{database}'

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Initialize database with app
db.init_app(app)

# Flask-Login configuration
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- DATABASE OLUŞTURMA ---
def init_db():
    conn = sqlite3.connect('sera.db')
    c = conn.cursor()

    # Üretim Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS uretim (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sera_adi TEXT NOT NULL,
        urun_adi TEXT NOT NULL,
        ekim_tarihi TEXT NOT NULL,
        hasat_tarihi TEXT,
        durum TEXT DEFAULT 'Ekim Yapıldı',
        alan REAL,
        beklenen_verim REAL,
        gercek_verim REAL,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    # Stok Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS stok (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        malzeme_adi TEXT NOT NULL,
        kategori TEXT,
        miktar REAL NOT NULL,
        birim TEXT,
        tarih TEXT NOT NULL,
        depo TEXT,
        min_stok REAL DEFAULT 0,
        maliyet REAL DEFAULT 0,
        notlar TEXT,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    # Personel Tablosu (Çalışan Bilgileri)
    c.execute('''CREATE TABLE IF NOT EXISTS personel (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_adi TEXT NOT NULL,
        pozisyon TEXT,
        aylik_maas REAL DEFAULT 0,
        ise_baslama_tarihi TEXT,
        aktif INTEGER DEFAULT 1,
        telefon TEXT,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    # Devam/Yoklama Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS devam (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL,
        tarih TEXT NOT NULL,
        durum TEXT NOT NULL, -- 'Geldi', 'Gelmedi', 'İzinli', 'Rapor'
        giris_saati TEXT,
        cikis_saati TEXT,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (personel_id) REFERENCES personel (id),
        UNIQUE(personel_id, tarih)
    )''')

    # Görev Tablosu (İşçilik yerine)
    c.execute('''CREATE TABLE IF NOT EXISTS gorevler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL,
        gorev TEXT NOT NULL,
        tarih TEXT NOT NULL,
        sera_adi TEXT,
        durum TEXT DEFAULT 'Tamamlandı',
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (personel_id) REFERENCES personel (id)
    )''')

    # Sulama Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS sulama (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        saat TEXT, -- Sulama saati (HH:MM format)
        sera_adi TEXT NOT NULL,
        sulama_turu TEXT NOT NULL, -- 'Normal', 'Gübreli', 'İlaçlı'
        miktar REAL NOT NULL, -- Litre
        gubre_kimyasal TEXT, -- Kullanılan gübre/kimyasal adı
        gubre_miktari REAL DEFAULT 0, -- Kullanılan gübre/kimyasal miktarı
        personel TEXT NOT NULL,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Add hour field to existing tables if it doesn't exist
    try:
        c.execute("ALTER TABLE sulama ADD COLUMN saat TEXT")
    except sqlite3.OperationalError:
        # Column already exists
        pass

    # Gübreleme Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS gubreleme (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tarih TEXT NOT NULL,
        sera_adi TEXT NOT NULL,
        gubre_adi TEXT NOT NULL,
        gubre_miktari REAL NOT NULL,
        uygulama_sekli TEXT, -- 'Yaprak', 'Kök', 'Karışım'
        personel TEXT NOT NULL,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    # Hasat Tablosu
    c.execute('''CREATE TABLE IF NOT EXISTS hasat (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        uretim_id INTEGER,
        hasat_tarihi TEXT NOT NULL,
        parsil_alan TEXT NOT NULL,
        hasat_miktari REAL NOT NULL,
        hasat_eden TEXT NOT NULL,
        teslim_edilen TEXT,
        kutu_sayisi REAL DEFAULT 0,
        palet_sayisi REAL DEFAULT 0,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (uretim_id) REFERENCES uretim (id)
    )''')

    # Faturalar Tablosu (Finans Modülü)
    c.execute('''CREATE TABLE IF NOT EXISTS faturalar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hasat_id INTEGER NOT NULL,
        firma_adi TEXT NOT NULL,
        hasat_tarihi TEXT NOT NULL,
        miktar REAL NOT NULL, -- kg
        birim_fiyat REAL, -- TL/kg (finans ekibi tarafından belirlenecek)
        toplam_tutar REAL, -- miktar * birim_fiyat
        durum TEXT DEFAULT 'Beklemede', -- 'Beklemede', 'Fiyatlandırıldı', 'Faturalandı', 'Tahsil Edildi'
        fiyat_belirleme_tarihi TEXT,
        fatura_tarihi TEXT,
        tahsil_tarihi TEXT,
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (hasat_id) REFERENCES hasat (id)
    )''')

    # Maliyetler Tablosu (Operasyon Giderleri Takibi)
    c.execute('''CREATE TABLE IF NOT EXISTS maliyetler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        maliyet_turu TEXT NOT NULL, -- 'Gübre', 'Kutu', 'Palet', 'İşçilik', 'Elektrik', 'Su', 'Diğer'
        tarih TEXT NOT NULL,
        aciklama TEXT NOT NULL,
        miktar REAL, -- Miktar (kg, adet, saat vs.)
        birim TEXT, -- kg, adet, saat, kWh vs.
        birim_fiyat REAL NOT NULL, -- TL/birim
        toplam_tutar REAL NOT NULL, -- miktar * birim_fiyat
        sera_adi TEXT, -- Hangi sera ile ilişkili (opsiyonel)
        uretim_id INTEGER, -- Hangi üretim dönemine ait (opsiyonel)
        hasat_id INTEGER, -- Hangi hasatla ilişkili (opsiyonel)
        tedarikci TEXT, -- Tedarikci bilgisi
        fatura_no TEXT, -- Fatura numarası
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (uretim_id) REFERENCES uretim (id),
        FOREIGN KEY (hasat_id) REFERENCES hasat (id)
    )''')

    # Kar/Zarar Tablosu (Hesaplanan Karlılık)
    c.execute('''CREATE TABLE IF NOT EXISTS kar_zarar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hesaplama_tarihi TEXT NOT NULL,
        donem_baslangic TEXT NOT NULL, -- Dönem başlangıç tarihi
        donem_bitis TEXT NOT NULL, -- Dönem bitiş tarihi
        sera_adi TEXT, -- Hangi sera (opsiyonel, tüm seralar için boş bırakılabilir)
        uretim_id INTEGER, -- Hangi üretim dönemine ait (opsiyonel)
        toplam_gelir REAL NOT NULL, -- Faturalardan gelen toplam gelir
        toplam_gider REAL NOT NULL, -- Maliyetlerden toplam gider
        net_kar_zarar REAL NOT NULL, -- toplam_gelir - toplam_gider
        karlillik_orani REAL, -- net_kar_zarar / toplam_gelir * 100
        hasat_miktari REAL, -- Toplam hasat miktarı (kg)
        kg_basina_maliyet REAL, -- toplam_gider / hasat_miktari
        kg_basina_gelir REAL, -- toplam_gelir / hasat_miktari
        kg_basina_kar REAL, -- net_kar_zarar / hasat_miktari
        notlar TEXT,
        olusturma_tarihi TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by TEXT,
        modified_by TEXT,
        modified_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (uretim_id) REFERENCES uretim (id)
    )''')

    # Mevcut tablolara audit alanlarını ekle (migration)
    try:
        # Check if audit columns exist, if not add them
        tables_to_migrate = ['uretim', 'stok', 'personel', 'devam', 'gorevler', 'hasat', 'sulama', 'gubreleme', 'faturalar', 'maliyetler', 'kar_zarar']
        
        for table in tables_to_migrate:
            # Check if created_by column exists
            c.execute(f"PRAGMA table_info({table})")
            columns = [row[1] for row in c.fetchall()]
            
            if 'created_by' not in columns:
                c.execute(f"ALTER TABLE {table} ADD COLUMN created_by TEXT")
            if 'modified_by' not in columns:
                c.execute(f"ALTER TABLE {table} ADD COLUMN modified_by TEXT")
            if 'modified_at' not in columns:
                c.execute(f"ALTER TABLE {table} ADD COLUMN modified_at TEXT")
                # Set initial values for existing records
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                c.execute(f"UPDATE {table} SET modified_at = ? WHERE modified_at IS NULL", (current_time,))
                
        # Add new columns to hasat table for box/pallet integration
        c.execute("PRAGMA table_info(hasat)")
        hasat_columns = [row[1] for row in c.fetchall()]
        if 'kutu_sayisi' not in hasat_columns:
            c.execute("ALTER TABLE hasat ADD COLUMN kutu_sayisi REAL DEFAULT 0")
        if 'palet_sayisi' not in hasat_columns:
            c.execute("ALTER TABLE hasat ADD COLUMN palet_sayisi REAL DEFAULT 0")
                
    except Exception as e:
        print(f"Migration error: {e}")
    
    conn.commit()
    conn.close()

def get_db_connection():
    conn = sqlite3.connect('sera.db')
    conn.row_factory = sqlite3.Row
    return conn

# --- AUDIT HELPER FUNCTIONS ---
def get_current_username():
    """Get current user's username for audit tracking"""
    if current_user.is_authenticated:
        return current_user.kullanici_adi
    return 'sistem'

def add_audit_fields_for_create():
    """Get audit fields for record creation"""
    username = get_current_username()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return {
        'created_by': username,
        'modified_by': username,
        'modified_at': timestamp
    }

def add_audit_fields_for_update():
    """Get audit fields for record update"""
    username = get_current_username()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return {
        'modified_by': username,
        'modified_at': timestamp
    }

# --- STOCK DEDUCTION HELPER FUNCTIONS ---
def deduct_stock_item(conn, kategori, miktar, operation_description=""):
    """
    Automatically deduct stock items by category
    Returns True if successful, False if insufficient stock
    """
    try:
        # Find the first available item in the category with sufficient stock
        item = conn.execute("""
            SELECT id, malzeme_adi, miktar, birim 
            FROM stok 
            WHERE kategori = ? AND miktar >= ? 
            ORDER BY tarih ASC 
            LIMIT 1
        """, (kategori, miktar)).fetchone()
        
        if not item:
            return False, f"Yetersiz {kategori.lower()} stoku (gerekli: {miktar})"
        
        # Deduct the quantity
        audit_fields = add_audit_fields_for_update()
        new_quantity = item['miktar'] - miktar
        
        conn.execute("""
            UPDATE stok 
            SET miktar = ?, modified_by = ?, modified_at = ?, 
                notlar = COALESCE(notlar, '') || CASE 
                    WHEN COALESCE(notlar, '') = '' THEN ? 
                    ELSE '; ' || ? 
                END
            WHERE id = ?
        """, (new_quantity, audit_fields['modified_by'], audit_fields['modified_at'],
              operation_description, operation_description, item['id']))
        
        return True, f"{miktar} {item['birim']} {item['malzeme_adi']} stoktan düşüldü"
        
    except Exception as e:
        return False, f"Stok düşme hatası: {str(e)}"

def batch_deduct_stock(conn, deductions):
    """
    Batch deduct multiple stock items
    deductions: list of (kategori, miktar, operation_description) tuples
    """
    results = []
    for kategori, miktar, operation_description in deductions:
        if miktar > 0:  # Only process if quantity > 0
            success, message = deduct_stock_item(conn, kategori, miktar, operation_description)
            results.append((success, message))
            if not success:
                return False, results  # Stop on first failure
    return True, results

# Initialize databases on startup
init_db()

# Create PostgreSQL tables and default admin user
with app.app_context():
    db.create_all()
    
    # Create default admin user if not exists
    admin_user = User.query.filter_by(kullanici_adi='admin').first()
    if not admin_user:
        admin_user = User(
            kullanici_adi='admin',
            email='admin@sera.com',
            ad_soyad='Sistem Yöneticisi',
            rol='admin'
        )
        admin_user.set_password('admin123')
        db.session.add(admin_user)
        db.session.commit()
        print("Default admin user created: admin / admin123")

# --- YARDIMCI FONKSİYONLAR ---
def get_dashboard_stats():
    conn = get_db_connection()
    
    # Aktif üretim sayısı
    aktif_uretim = conn.execute(
        "SELECT COUNT(*) as count FROM uretim WHERE durum IN ('Ekim Yapıldı', 'Büyüme Döneminde', 'Çiçeklenme')"
    ).fetchone()['count']
    
    # Düşük stok uyarıları
    dusuk_stok = conn.execute(
        "SELECT COUNT(*) as count FROM stok WHERE miktar <= min_stok AND min_stok > 0"
    ).fetchone()['count']
    
    # Bu ayki toplam personel maliyeti (sadece aktif personel)
    bu_ay = datetime.now().strftime('%Y-%m')
    aylik_personel = conn.execute(
        "SELECT COALESCE(SUM(aylik_maas), 0) as total FROM personel WHERE aktif = 1"
    ).fetchone()['total']
    
    # Toplam sera sayısı
    sera_sayisi = conn.execute(
        "SELECT COUNT(DISTINCT sera_adi) as count FROM uretim"
    ).fetchone()['count']
    
    # Bu ayki toplam hasat
    bu_ay_hasat = conn.execute(
        "SELECT COALESCE(SUM(hasat_miktari), 0) as total FROM hasat WHERE hasat_tarihi LIKE ?",
        (f"{bu_ay}%",)
    ).fetchone()['total']
    
    conn.close()
    
    return {
        'aktif_uretim': aktif_uretim,
        'dusuk_stok': dusuk_stok,
        'aylik_personel': aylik_personel,
        'sera_sayisi': sera_sayisi,
        'bu_ay_hasat': bu_ay_hasat
    }

# --- YARDIMCI DECORATORLER ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('Bu sayfaya erişmek için admin yetkisine sahip olmalısınız.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# --- ROTALAR ---

@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == "POST":
        kullanici_adi = request.form['kullanici_adi'].strip()
        password = request.form['password']
        remember = bool(request.form.get('remember'))
        
        # Login attempt kaydı
        attempt = LoginAttempt(
            kullanici_adi=kullanici_adi,
            ip_adresi=request.remote_addr
        )
        
        user = User.query.filter_by(kullanici_adi=kullanici_adi, aktif=True).first()
        
        if user and user.check_password(password):
            login_user(user, remember=remember)
            user.son_giris = datetime.utcnow()
            attempt.basarili = True
            db.session.add(attempt)
            db.session.commit()
            
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            
            flash(f'Hoş geldiniz, {user.ad_soyad or user.kullanici_adi}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            attempt.basarili = False
            db.session.add(attempt)
            db.session.commit()
            flash('Kullanıcı adı veya şifre hatalı!', 'error')
    
    return render_template('login.html')

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash('Başarıyla çıkış yaptınız.', 'info')
    return redirect(url_for('login'))

@app.route("/admin")
@login_required
@admin_required
def admin_panel():
    # Kullanıcı listesi
    users = User.query.order_by(User.olusturma_tarihi.desc()).all()
    
    # Son giriş denemeleri
    recent_attempts = LoginAttempt.query.order_by(LoginAttempt.deneme_tarihi.desc()).limit(20).all()
    
    # İstatistikler
    stats = {
        'total_users': User.query.count(),
        'active_users': User.query.filter_by(aktif=True).count(),
        'admin_users': User.query.filter_by(rol='admin').count(),
        'failed_attempts_today': LoginAttempt.query.filter(
            LoginAttempt.deneme_tarihi >= datetime.utcnow().date(),
            LoginAttempt.basarili == False
        ).count()
    }
    
    return render_template('admin.html', users=users, recent_attempts=recent_attempts, stats=stats)

@app.route("/admin/user/add", methods=["POST"])
@login_required
@admin_required
def admin_add_user():
    try:
        kullanici_adi = request.form['kullanici_adi'].strip()
        password = request.form['password']
        ad_soyad = request.form['ad_soyad'].strip()
        email = request.form.get('email', '').strip()
        rol = request.form['rol']
        
        # Kullanıcı adı kontrolü
        if User.query.filter_by(kullanici_adi=kullanici_adi).first():
            flash('Bu kullanıcı adı zaten kullanımda!', 'error')
            return redirect(url_for('admin_panel'))
        
        # Email kontrolü
        if email and User.query.filter_by(email=email).first():
            flash('Bu e-posta adresi zaten kullanımda!', 'error')
            return redirect(url_for('admin_panel'))
        
        # Yeni kullanıcı oluştur
        new_user = User(
            kullanici_adi=kullanici_adi,
            ad_soyad=ad_soyad,
            email=email if email else None,
            rol=rol
        )
        new_user.set_password(password)
        
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'Kullanıcı "{kullanici_adi}" başarıyla oluşturuldu!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Kullanıcı oluşturulurken hata: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))

@app.route("/admin/user/toggle/<int:user_id>")
@login_required
@admin_required
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        flash('Kendi hesabınızı devre dışı bırakamazsınız!', 'error')
    else:
        user.aktif = not user.aktif
        db.session.commit()
        
        status = 'aktif' if user.aktif else 'pasif'
        flash(f'Kullanıcı "{user.kullanici_adi}" {status} duruma getirildi.', 'success')
    
    return redirect(url_for('admin_panel'))

@app.route("/admin/user/reset-password/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = User.query.get_or_404(user_id)
    new_password = request.form['new_password']
    
    user.set_password(new_password)
    db.session.commit()
    
    flash(f'Kullanıcı "{user.kullanici_adi}" şifresi başarıyla güncellendi!', 'success')
    return redirect(url_for('admin_panel'))

@app.route("/dashboard")
@login_required
def dashboard():
    stats = get_dashboard_stats()
    
    conn = get_db_connection()
    
    # Son aktiviteler
    son_uretim = conn.execute(
        "SELECT * FROM uretim ORDER BY olusturma_tarihi DESC LIMIT 5"
    ).fetchall()
    
    son_stok = conn.execute(
        "SELECT * FROM stok ORDER BY tarih DESC LIMIT 5"
    ).fetchall()
    
    son_gorevler = conn.execute(
        "SELECT g.*, p.personel_adi FROM gorevler g LEFT JOIN personel p ON g.personel_id = p.id ORDER BY g.tarih DESC LIMIT 5"
    ).fetchall()
    
    conn.close()
    
    return render_template('dashboard.html', 
                         stats=stats,
                         son_uretim=son_uretim,
                         son_stok=son_stok,
                         son_gorevler=son_gorevler)

@app.route("/uretim", methods=["GET", "POST"])
@login_required
def uretim():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            sera_adi = request.form['sera_adi'].strip()
            urun_adi = request.form['urun_adi'].strip()
            ekim_tarihi = request.form['ekim_tarihi']
            hasat_tarihi = request.form.get('hasat_tarihi') or None
            alan = float(request.form.get('alan', 0) or 0)
            beklenen_verim = float(request.form.get('beklenen_verim', 0) or 0)
            notlar = request.form.get('notlar', '').strip()
            
            if not sera_adi or not urun_adi or not ekim_tarihi:
                flash('Sera adı, ürün adı ve ekim tarihi zorunludur!', 'error')
            else:
                audit_fields = add_audit_fields_for_create()
                conn.execute("""
                    INSERT INTO uretim (sera_adi, urun_adi, ekim_tarihi, hasat_tarihi, 
                                      alan, beklenen_verim, notlar, durum, created_by, modified_by, modified_at) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (sera_adi, urun_adi, ekim_tarihi, hasat_tarihi, alan, beklenen_verim, notlar, "Ekim Yapıldı",
                      audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                conn.commit()
                flash('Üretim kaydı başarıyla eklendi!', 'success')
                return redirect(url_for('uretim'))
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    uretimler = conn.execute("SELECT * FROM uretim ORDER BY ekim_tarihi DESC").fetchall()
    conn.close()
    
    return render_template("uretim.html", uretimler=uretimler)

@app.route("/uretim/guncelle/<int:id>", methods=["POST"])
@login_required
def uretim_guncelle(id):
    conn = get_db_connection()
    
    try:
        durum = request.form.get('durum')
        gercek_verim = request.form.get('gercek_verim')
        notlar = request.form.get('notlar', '')
        
        if gercek_verim:
            gercek_verim = float(gercek_verim)
        else:
            gercek_verim = None
            
        audit_fields = add_audit_fields_for_update()
        conn.execute("""
            UPDATE uretim 
            SET durum = ?, gercek_verim = ?, notlar = ?, modified_by = ?, modified_at = ?
            WHERE id = ?
        """, (durum, gercek_verim, notlar, audit_fields['modified_by'], audit_fields['modified_at'], id))
        conn.commit()
        flash('Üretim kaydı güncellendi!', 'success')
    except Exception as e:
        flash(f'Güncelleme sırasında hata: {str(e)}', 'error')
    
    conn.close()
    return redirect(url_for('uretim'))

@app.route("/stok", methods=["GET", "POST"])
@login_required
def stok():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            malzeme_adi = request.form['malzeme_adi'].strip()
            kategori = request.form.get('kategori', '').strip()
            miktar = float(request.form['miktar'])
            birim = request.form.get('birim', '').strip()
            depo = request.form.get('depo', '').strip()
            min_stok = float(request.form.get('min_stok', 0) or 0)
            maliyet = float(request.form.get('maliyet', 0) or 0)
            notlar = request.form.get('notlar', '').strip()
            islem_turu = request.form.get('islem_turu', 'ekle')
            
            if not malzeme_adi:
                flash('Malzeme adı zorunludur!', 'error')
            else:
                # Mevcut stok kontrolü
                mevcut = conn.execute(
                    "SELECT * FROM stok WHERE malzeme_adi = ? AND depo = ?", 
                    (malzeme_adi, depo)
                ).fetchone()
                
                if mevcut and islem_turu == 'cikar':
                    # Stok çıkarma
                    yeni_miktar = mevcut['miktar'] - miktar
                    if yeni_miktar < 0:
                        flash('Yetersiz stok! Mevcut miktar: {}'.format(mevcut['miktar']), 'error')
                    else:
                        audit_fields = add_audit_fields_for_update()
                        conn.execute(
                            "UPDATE stok SET miktar = ?, modified_by = ?, modified_at = ? WHERE id = ?",
                            (yeni_miktar, audit_fields['modified_by'], audit_fields['modified_at'], mevcut['id'])
                        )
                        conn.commit()
                        flash('Stok çıkarıldı!', 'success')
                elif mevcut and islem_turu == 'ekle':
                    # Mevcut stoka ekleme
                    yeni_miktar = mevcut['miktar'] + miktar
                    audit_fields = add_audit_fields_for_update()
                    conn.execute(
                        "UPDATE stok SET miktar = ?, maliyet = ?, min_stok = ?, notlar = ?, modified_by = ?, modified_at = ? WHERE id = ?",
                        (yeni_miktar, maliyet, min_stok, notlar, audit_fields['modified_by'], audit_fields['modified_at'], mevcut['id'])
                    )
                    conn.commit()
                    flash('Stok güncellendi!', 'success')
                else:
                    # Yeni stok kaydı
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO stok (malzeme_adi, kategori, miktar, birim, tarih, 
                                        depo, min_stok, maliyet, notlar, created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (malzeme_adi, kategori, miktar, birim, datetime.now().strftime('%Y-%m-%d'),
                          depo, min_stok, maliyet, notlar, audit_fields['created_by'], 
                          audit_fields['modified_by'], audit_fields['modified_at']))
                    conn.commit()
                    flash('Yeni stok kaydı eklendi!', 'success')
                
                return redirect(url_for('stok'))
                
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    stoklar = conn.execute("SELECT * FROM stok ORDER BY malzeme_adi").fetchall()
    
    # Düşük stok uyarıları
    dusuk_stoklar = conn.execute(
        "SELECT * FROM stok WHERE miktar <= min_stok AND min_stok > 0"
    ).fetchall()
    
    conn.close()
    
    return render_template("stok.html", stoklar=stoklar, dusuk_stoklar=dusuk_stoklar)

@app.route("/personel", methods=["GET", "POST"])
@login_required
def personel():
    conn = get_db_connection()
    
    if request.method == "POST":
        action = request.form.get('action', '')
        
        if action == 'personel_ekle':
            try:
                personel_adi = request.form['personel_adi'].strip()
                pozisyon = request.form.get('pozisyon', '').strip()
                aylik_maas = float(request.form.get('aylik_maas', 0) or 0)
                ise_baslama_tarihi = request.form.get('ise_baslama_tarihi')
                telefon = request.form.get('telefon', '').strip()
                notlar = request.form.get('notlar', '').strip()
                
                if not personel_adi:
                    flash('Personel adı zorunludur!', 'error')
                else:
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO personel (personel_adi, pozisyon, aylik_maas, ise_baslama_tarihi, telefon, notlar,
                                            created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (personel_adi, pozisyon, aylik_maas, ise_baslama_tarihi, telefon, notlar,
                          audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                    conn.commit()
                    flash('Personel kaydı başarıyla eklendi!', 'success')
                    return redirect(url_for('personel'))
                    
            except ValueError:
                flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
            except Exception as e:
                flash(f'Bir hata oluştu: {str(e)}', 'error')
        
        elif action == 'gorev_ekle':
            try:
                personel_id = request.form['personel_id']
                gorev = request.form['gorev'].strip()
                tarih = request.form['tarih']
                sera_adi = request.form.get('sera_adi', '').strip()
                notlar = request.form.get('notlar', '').strip()
                
                if not personel_id or not gorev or not tarih:
                    flash('Personel, görev ve tarih zorunludur!', 'error')
                else:
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO gorevler (personel_id, gorev, tarih, sera_adi, notlar, created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (personel_id, gorev, tarih, sera_adi, notlar,
                          audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                    conn.commit()
                    flash('Görev kaydı başarıyla eklendi!', 'success')
                    return redirect(url_for('personel'))
                    
            except Exception as e:
                flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    # Personel listesi
    personeller = conn.execute("SELECT * FROM personel ORDER BY aktif DESC, personel_adi").fetchall()
    
    # Görev kayıtları
    gorevler = conn.execute("""
        SELECT g.*, p.personel_adi 
        FROM gorevler g 
        LEFT JOIN personel p ON g.personel_id = p.id 
        ORDER BY g.tarih DESC
    """).fetchall()
    
    # Bu ayki toplam maliyet
    bu_ay_toplam_maas = conn.execute(
        "SELECT COALESCE(SUM(aylik_maas), 0) as total FROM personel WHERE aktif = 1"
    ).fetchone()['total']
    
    # Personel istatistikleri
    bu_ay = datetime.now().strftime('%Y-%m')
    personel_istatistik = conn.execute("""
        SELECT p.personel_adi, p.aylik_maas,
               COUNT(g.id) as gorev_sayisi
        FROM personel p 
        LEFT JOIN gorevler g ON p.id = g.personel_id AND g.tarih LIKE ?
        WHERE p.aktif = 1
        GROUP BY p.id, p.personel_adi, p.aylik_maas
        ORDER BY p.personel_adi
    """, (f"{bu_ay}%",)).fetchall()
    
    conn.close()
    
    return render_template("personel.html", 
                         personeller=personeller,
                         gorevler=gorevler,
                         bu_ay_toplam_maas=bu_ay_toplam_maas,
                         personel_istatistik=personel_istatistik)

@app.route("/devam", methods=["GET", "POST"])
@login_required
def devam():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            tarih = request.form['tarih']
            
            # Tüm aktif personel için devam durumunu kaydet
            personeller = conn.execute("SELECT * FROM personel WHERE aktif = 1").fetchall()
            
            for personel in personeller:
                durum = request.form.get(f'durum_{personel["id"]}', 'Gelmedi')
                giris_saati = request.form.get(f'giris_{personel["id"]}', '').strip()
                cikis_saati = request.form.get(f'cikis_{personel["id"]}', '').strip()
                notlar = request.form.get(f'notlar_{personel["id"]}', '').strip()
                
                # Mevcut kaydı kontrol et
                mevcut = conn.execute(
                    "SELECT * FROM devam WHERE personel_id = ? AND tarih = ?",
                    (personel['id'], tarih)
                ).fetchone()
                
                if mevcut:
                    # Güncelle
                    audit_fields = add_audit_fields_for_update()
                    conn.execute("""
                        UPDATE devam 
                        SET durum = ?, giris_saati = ?, cikis_saati = ?, notlar = ?, modified_by = ?, modified_at = ?
                        WHERE personel_id = ? AND tarih = ?
                    """, (durum, giris_saati, cikis_saati, notlar, audit_fields['modified_by'], audit_fields['modified_at'], personel['id'], tarih))
                else:
                    # Yeni kayıt
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO devam (personel_id, tarih, durum, giris_saati, cikis_saati, notlar,
                                         created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (personel['id'], tarih, durum, giris_saati, cikis_saati, notlar,
                          audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
            
            conn.commit()
            flash('Devam durumu başarıyla kaydedildi!', 'success')
            return redirect(url_for('devam'))
            
        except Exception as e:
            flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    # Bugünün tarihi
    bugun = datetime.now().strftime('%Y-%m-%d')
    
    # Aktif personel listesi
    personeller = conn.execute("SELECT * FROM personel WHERE aktif = 1 ORDER BY personel_adi").fetchall()
    
    # Bugünkü devam durumu
    bugun_devam = conn.execute("""
        SELECT d.*, p.personel_adi 
        FROM devam d 
        LEFT JOIN personel p ON d.personel_id = p.id 
        WHERE d.tarih = ?
    """, (bugun,)).fetchall()
    
    # Son 7 günün devam istatistikleri
    devam_istatistik = conn.execute("""
        SELECT d.tarih,
               COUNT(*) as toplam_personel,
               SUM(CASE WHEN d.durum = 'Geldi' THEN 1 ELSE 0 END) as gelenler,
               SUM(CASE WHEN d.durum = 'Gelmedi' THEN 1 ELSE 0 END) as gelmeyenler,
               SUM(CASE WHEN d.durum = 'İzinli' THEN 1 ELSE 0 END) as izinliler
        FROM devam d 
        WHERE d.tarih >= date('now', '-7 days')
        GROUP BY d.tarih
        ORDER BY d.tarih DESC
    """).fetchall()
    
    conn.close()
    
    return render_template("devam.html", 
                         personeller=personeller,
                         bugun_devam=bugun_devam,
                         devam_istatistik=devam_istatistik,
                         bugun=bugun)

@app.route("/devam/excel-export")
@login_required
def devam_excel_export():
    conn = get_db_connection()
    
    # Tarih aralığı parametreleri
    baslangic_tarihi = request.args.get('baslangic', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    bitis_tarihi = request.args.get('bitis', datetime.now().strftime('%Y-%m-%d'))
    
    # Devam kayıtlarını al
    devam_kayitlari = conn.execute("""
        SELECT d.tarih, p.personel_adi, p.pozisyon, d.durum, d.giris_saati, d.cikis_saati, d.notlar
        FROM devam d
        LEFT JOIN personel p ON d.personel_id = p.id
        WHERE d.tarih BETWEEN ? AND ?
        ORDER BY d.tarih DESC, p.personel_adi
    """, (baslangic_tarihi, bitis_tarihi)).fetchall()
    
    # Personel bazında özet istatistikler
    personel_ozet = conn.execute("""
        SELECT p.personel_adi, p.pozisyon,
               COUNT(*) as toplam_gun,
               SUM(CASE WHEN d.durum = 'Geldi' THEN 1 ELSE 0 END) as gelme_sayisi,
               SUM(CASE WHEN d.durum = 'Gelmedi' THEN 1 ELSE 0 END) as gelmeme_sayisi,
               SUM(CASE WHEN d.durum = 'İzinli' THEN 1 ELSE 0 END) as izin_sayisi,
               SUM(CASE WHEN d.durum = 'Rapor' THEN 1 ELSE 0 END) as rapor_sayisi
        FROM devam d
        LEFT JOIN personel p ON d.personel_id = p.id
        WHERE d.tarih BETWEEN ? AND ?
        GROUP BY p.personel_adi, p.pozisyon
        ORDER BY p.personel_adi
    """, (baslangic_tarihi, bitis_tarihi)).fetchall()
    
    conn.close()
    
    # Excel dosyası oluştur
    wb = Workbook()
    
    # İlk sayfa: Günlük devam kayıtları
    ws1 = wb.active
    ws1.title = "Günlük Devam Kayıtları"
    
    # Başlık stilleri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Başlıklar
    headers1 = ["Tarih", "Personel Adı", "Pozisyon", "Durum", "Giriş Saati", "Çıkış Saati", "Notlar"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Veri satırları
    for row, devam in enumerate(devam_kayitlari, 2):
        ws1.cell(row=row, column=1, value=devam['tarih'])
        ws1.cell(row=row, column=2, value=devam['personel_adi'])
        ws1.cell(row=row, column=3, value=devam['pozisyon'] or '-')
        ws1.cell(row=row, column=4, value=devam['durum'])
        ws1.cell(row=row, column=5, value=devam['giris_saati'] or '-')
        ws1.cell(row=row, column=6, value=devam['cikis_saati'] or '-')
        ws1.cell(row=row, column=7, value=devam['notlar'] or '-')
    
    # Sütun genişliklerini ayarla
    column_widths1 = [12, 20, 15, 12, 12, 12, 30]
    for col, width in enumerate(column_widths1, 1):
        ws1.column_dimensions[chr(64 + col)].width = width
    
    # İkinci sayfa: Personel özet istatistikleri
    ws2 = wb.create_sheet(title="Personel Özet İstatistikleri")
    
    # Başlıklar
    headers2 = ["Personel Adı", "Pozisyon", "Toplam Gün", "Gelme", "Gelmeme", "İzin", "Rapor", "Devam Oranı (%)"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Veri satırları
    for row, ozet in enumerate(personel_ozet, 2):
        devam_orani = (ozet['gelme_sayisi'] / ozet['toplam_gun'] * 100) if ozet['toplam_gun'] > 0 else 0
        
        ws2.cell(row=row, column=1, value=ozet['personel_adi'])
        ws2.cell(row=row, column=2, value=ozet['pozisyon'] or '-')
        ws2.cell(row=row, column=3, value=ozet['toplam_gun'])
        ws2.cell(row=row, column=4, value=ozet['gelme_sayisi'])
        ws2.cell(row=row, column=5, value=ozet['gelmeme_sayisi'])
        ws2.cell(row=row, column=6, value=ozet['izin_sayisi'])
        ws2.cell(row=row, column=7, value=ozet['rapor_sayisi'])
        ws2.cell(row=row, column=8, value=f"{devam_orani:.1f}%")
    
    # Sütun genişliklerini ayarla
    column_widths2 = [20, 15, 12, 8, 10, 8, 8, 15]
    for col, width in enumerate(column_widths2, 1):
        ws2.column_dimensions[chr(64 + col)].width = width
    
    # Excel dosyasını belleğe kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adı
    filename = f"devam_raporu_{baslangic_tarihi}_{bitis_tarihi}.xlsx"
    
    # Response oluştur
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    return response

@app.route("/hasat", methods=["GET", "POST"])
@login_required
def hasat():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            uretim_id = request.form.get('uretim_id')
            hasat_tarihi = request.form['hasat_tarihi']
            parsil_alan = request.form['parsil_alan'].strip()
            hasat_miktari = float(request.form['hasat_miktari'])
            hasat_eden = request.form['hasat_eden'].strip()
            teslim_edilen = request.form.get('teslim_edilen', '').strip()
            kutu_sayisi = float(request.form.get('kutu_sayisi', 0) or 0)
            palet_sayisi = float(request.form.get('palet_sayisi', 0) or 0)
            notlar = request.form.get('notlar', '').strip()
            
            if not parsil_alan or not hasat_eden or not hasat_tarihi:
                flash('Hasat tarihi, parsel/alan ve hasat eden kişi zorunludur!', 'error')
            else:
                # Prepare stock deductions
                deductions = []
                if kutu_sayisi > 0:
                    deductions.append(('Kutu', kutu_sayisi, f'Hasat kullanımı - {hasat_tarihi} - {parsil_alan}'))
                if palet_sayisi > 0:
                    deductions.append(('Palet', palet_sayisi, f'Hasat kullanımı - {hasat_tarihi} - {parsil_alan}'))
                
                # Check stock availability and deduct if possible
                stock_success = True
                stock_messages = []
                
                if deductions:
                    stock_success, stock_results = batch_deduct_stock(conn, deductions)
                    stock_messages = [result[1] for result in stock_results]
                
                if stock_success:
                    audit_fields = add_audit_fields_for_create()
                    
                    # Insert hasat record
                    hasat_cursor = conn.execute("""
                        INSERT INTO hasat (uretim_id, hasat_tarihi, parsil_alan, hasat_miktari, 
                                         hasat_eden, teslim_edilen, kutu_sayisi, palet_sayisi, notlar, 
                                         created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (uretim_id, hasat_tarihi, parsil_alan, hasat_miktari, 
                          hasat_eden, teslim_edilen, kutu_sayisi, palet_sayisi, notlar, 
                          audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                    
                    hasat_id = hasat_cursor.lastrowid
                    
                    # Otomatik fatura kaydı oluştur (eğer teslim edilen firma belirtilmişse)
                    if teslim_edilen:
                        try:
                            conn.execute("""
                                INSERT INTO faturalar (hasat_id, firma_adi, hasat_tarihi, miktar, 
                                                     durum, created_by, modified_by, modified_at) 
                                VALUES (?, ?, ?, ?, 'Beklemede', ?, ?, ?)
                            """, (hasat_id, teslim_edilen, hasat_tarihi, hasat_miktari,
                                  audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                        except Exception as e:
                            # Fatura kaydı başarısız olsa da hasat kaydı devam etsin
                            print(f"Fatura oluşturma hatası: {e}")
                    
                    conn.commit()
                    
                    success_msg = 'Hasat kaydı başarıyla eklendi!'
                    if teslim_edilen:
                        success_msg += f' Finans modülünde "{teslim_edilen}" için fatura kaydı oluşturuldu.'
                    if stock_messages:
                        success_msg += ' Stok güncellemeleri: ' + ', '.join(stock_messages)
                    flash(success_msg, 'success')
                    return redirect(url_for('hasat'))
                else:
                    conn.rollback()
                    error_msg = 'Stok yetersiz: ' + ', '.join([result[1] for result in stock_results if not result[0]])
                    flash(error_msg, 'error')
                
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    # Hasat kayıtları ve ilgili üretim bilgileri
    hasatlar = conn.execute("""
        SELECT h.*, u.sera_adi, u.urun_adi, u.ekim_tarihi
        FROM hasat h
        LEFT JOIN uretim u ON h.uretim_id = u.id
        ORDER BY h.hasat_tarihi DESC
    """).fetchall()
    
    # Aktif üretimler (hasat için)
    aktif_uretimler = conn.execute(
        "SELECT * FROM uretim WHERE durum != 'Hasat Edildi' ORDER BY sera_adi, urun_adi"
    ).fetchall()
    
    # Bu ayki hasat istatistikleri
    bu_ay = datetime.now().strftime('%Y-%m')
    bu_ay_toplam = conn.execute(
        "SELECT COALESCE(SUM(hasat_miktari), 0) as total FROM hasat WHERE hasat_tarihi LIKE ?",
        (f"{bu_ay}%",)
    ).fetchone()['total']
    
    # En çok hasat yapan kişiler
    hasat_eden_istatistik = conn.execute("""
        SELECT hasat_eden,
               COUNT(*) as hasat_sayisi,
               SUM(hasat_miktari) as toplam_miktar
        FROM hasat 
        WHERE hasat_tarihi LIKE ?
        GROUP BY hasat_eden
        ORDER BY toplam_miktar DESC
        LIMIT 5
    """, (f"{bu_ay}%",)).fetchall()
    
    conn.close()
    
    return render_template("hasat.html", 
                         hasatlar=hasatlar,
                         aktif_uretimler=aktif_uretimler,
                         bu_ay_toplam=bu_ay_toplam,
                         hasat_eden_istatistik=hasat_eden_istatistik)

@app.route("/sulama", methods=["GET", "POST"])
@login_required
def sulama():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            # Check if this is a bulk entry
            bulk_entry = request.form.get('bulk_entry', False)
            
            if bulk_entry:
                # Handle bulk entry
                tarih = request.form['bulk_tarih']
                sera_adi = request.form['bulk_sera_adi'].strip()
                personel = request.form['bulk_personel'].strip()
                
                # Process multiple irrigation entries
                entries_processed = 0
                total_deductions = []
                
                # Get all form fields for bulk entries
                i = 1
                while f'saat_{i}' in request.form:
                    saat = request.form.get(f'saat_{i}', '').strip()
                    sulama_turu = request.form.get(f'sulama_turu_{i}')
                    miktar = float(request.form.get(f'miktar_{i}', 0) or 0)
                    gubre_kimyasal = request.form.get(f'gubre_kimyasal_{i}', '').strip()
                    gubre_miktari = float(request.form.get(f'gubre_miktari_{i}', 0) or 0)
                    notlar = request.form.get(f'notlar_{i}', '').strip()
                    
                    if saat and sulama_turu and miktar > 0:
                        # Prepare stock deductions for this entry
                        if gubre_miktari > 0 and gubre_kimyasal:
                            if sulama_turu == 'Gübreli':
                                total_deductions.append(('Gübre', gubre_miktari, f'Sulama kullanımı - {tarih} {saat} - {sera_adi}'))
                            elif sulama_turu == 'İlaçlı':
                                total_deductions.append(('Pestisit', gubre_miktari, f'Sulama kullanımı - {tarih} {saat} - {sera_adi}'))
                                total_deductions.append(('Kimyasal', gubre_miktari, f'Sulama kullanımı - {tarih} {saat} - {sera_adi}'))
                        
                        entries_processed += 1
                    i += 1
                
                if entries_processed == 0:
                    flash('En az bir geçerli sulama kaydı girmelisiniz!', 'error')
                else:
                    # Process all stock deductions at once
                    stock_success = True
                    stock_messages = []
                    
                    if total_deductions:
                        stock_success, stock_results = batch_deduct_stock(conn, total_deductions)
                        stock_messages = [result[1] for result in stock_results if result[0]]
                    
                    if stock_success:
                        # Insert all entries
                        audit_fields = add_audit_fields_for_create()
                        i = 1
                        while f'saat_{i}' in request.form:
                            saat = request.form.get(f'saat_{i}', '').strip()
                            sulama_turu = request.form.get(f'sulama_turu_{i}')
                            miktar = float(request.form.get(f'miktar_{i}', 0) or 0)
                            gubre_kimyasal = request.form.get(f'gubre_kimyasal_{i}', '').strip()
                            gubre_miktari = float(request.form.get(f'gubre_miktari_{i}', 0) or 0)
                            notlar = request.form.get(f'notlar_{i}', '').strip()
                            
                            if saat and sulama_turu and miktar > 0:
                                conn.execute("""
                                    INSERT INTO sulama (tarih, saat, sera_adi, sulama_turu, miktar, gubre_kimyasal, 
                                                      gubre_miktari, personel, notlar, created_by, modified_by, modified_at) 
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, (tarih, saat, sera_adi, sulama_turu, miktar, gubre_kimyasal, gubre_miktari, 
                                      personel, notlar, audit_fields['created_by'], audit_fields['modified_by'], 
                                      audit_fields['modified_at']))
                            i += 1
                        
                        conn.commit()
                        success_msg = f'{entries_processed} sulama kaydı başarıyla eklendi!'
                        if stock_messages:
                            success_msg += ' Stok güncellemeleri: ' + ', '.join(stock_messages)
                        flash(success_msg, 'success')
                        return redirect(url_for('sulama'))
                    else:
                        conn.rollback()
                        error_msg = 'Stok yetersiz: ' + ', '.join([result[1] for result in stock_results if not result[0]])
                        flash(error_msg, 'error')
            else:
                # Handle single entry
                tarih = request.form['tarih']
                saat = request.form.get('saat', '').strip()
                sera_adi = request.form['sera_adi'].strip()
                sulama_turu = request.form['sulama_turu']
                miktar = float(request.form['miktar'])
                gubre_kimyasal = request.form.get('gubre_kimyasal', '').strip()
                gubre_miktari = float(request.form.get('gubre_miktari', 0) or 0)
                personel = request.form['personel'].strip()
                notlar = request.form.get('notlar', '').strip()
            
            if not sera_adi or not personel or not tarih or miktar <= 0:
                flash('Sera adı, personel, tarih ve miktar zorunludur!', 'error')
            else:
                # Prepare stock deductions
                deductions = []
                if gubre_miktari > 0 and gubre_kimyasal:
                    if sulama_turu == 'Gübreli':
                        deductions.append(('Gübre', gubre_miktari, f'Sulama kullanımı - {tarih} - {sera_adi}'))
                    elif sulama_turu == 'İlaçlı':
                        deductions.append(('Pestisit', gubre_miktari, f'Sulama kullanımı - {tarih} - {sera_adi}'))
                        deductions.append(('Kimyasal', gubre_miktari, f'Sulama kullanımı - {tarih} - {sera_adi}'))
                
                # Process stock deductions
                stock_success = True
                stock_messages = []
                
                if deductions:
                    stock_success, stock_results = batch_deduct_stock(conn, deductions)
                    stock_messages = [result[1] for result in stock_results if result[0]]
                
                if stock_success:
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO sulama (tarih, saat, sera_adi, sulama_turu, miktar, gubre_kimyasal, 
                                          gubre_miktari, personel, notlar, created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (tarih, saat, sera_adi, sulama_turu, miktar, gubre_kimyasal, gubre_miktari, 
                          personel, notlar, audit_fields['created_by'], audit_fields['modified_by'], 
                          audit_fields['modified_at']))
                    conn.commit()
                    
                    success_msg = 'Sulama kaydı başarıyla eklendi!'
                    if stock_messages:
                        success_msg += ' Stok güncellemeleri: ' + ', '.join(stock_messages)
                    flash(success_msg, 'success')
                    return redirect(url_for('sulama'))
                else:
                    conn.rollback()
                    error_msg = 'Stok yetersiz: ' + ', '.join([result[1] for result in stock_results if not result[0]])
                    flash(error_msg, 'error')
                
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Hata oluştu: {str(e)}', 'error')
            conn.rollback()
        finally:
            conn.close()
    
    # GET request - show irrigation records
    conn = get_db_connection()
    
    # Get irrigation records
    sulamalar = conn.execute("""
        SELECT * FROM sulama 
        ORDER BY tarih DESC, olusturma_tarihi DESC
    """).fetchall()
    
    # Get monthly statistics
    bu_ay = datetime.now().strftime('%Y-%m')
    bu_ay_sulama = conn.execute("""
        SELECT COALESCE(SUM(miktar), 0) as total 
        FROM sulama 
        WHERE tarih LIKE ?
    """, (f"{bu_ay}%",)).fetchone()['total']
    
    # Get fertilizer usage statistics
    gubre_istatistik = conn.execute("""
        SELECT sulama_turu, COUNT(*) as kayit_sayisi, COALESCE(SUM(gubre_miktari), 0) as toplam_miktar
        FROM sulama 
        WHERE tarih LIKE ? AND gubre_miktari > 0
        GROUP BY sulama_turu
    """, (f"{bu_ay}%",)).fetchall()
    
    # Get available fertilizers and chemicals from stock
    available_fertilizers = conn.execute("""
        SELECT DISTINCT malzeme_adi, miktar, birim
        FROM stok 
        WHERE kategori = 'Gübre' AND miktar > 0
        ORDER BY malzeme_adi
    """).fetchall()
    
    available_chemicals = conn.execute("""
        SELECT DISTINCT malzeme_adi, miktar, birim
        FROM stok 
        WHERE kategori IN ('Pestisit', 'Kimyasal') AND miktar > 0
        ORDER BY malzeme_adi
    """).fetchall()
    
    # Get daily summaries for the last 7 days
    daily_summaries = conn.execute("""
        SELECT tarih, 
               COUNT(*) as toplam_seans,
               COALESCE(SUM(miktar), 0) as toplam_su,
               COALESCE(GROUP_CONCAT(DISTINCT saat), '') as saatler,
               COALESCE(SUM(CASE WHEN sulama_turu = 'Normal' THEN 1 ELSE 0 END), 0) as normal_seans,
               COALESCE(SUM(CASE WHEN sulama_turu = 'Gübreli' THEN 1 ELSE 0 END), 0) as gubreli_seans,
               COALESCE(SUM(CASE WHEN sulama_turu = 'İlaçlı' THEN 1 ELSE 0 END), 0) as ilacli_seans
        FROM sulama 
        WHERE tarih >= date('now', '-7 days')
        GROUP BY tarih
        ORDER BY tarih DESC
    """).fetchall()
    
    conn.close()
    
    return render_template('sulama.html', 
                         sulamalar=sulamalar,
                         bu_ay_sulama=bu_ay_sulama,
                         gubre_istatistik=gubre_istatistik,
                         available_fertilizers=available_fertilizers,
                         available_chemicals=available_chemicals,
                         daily_summaries=daily_summaries)

@app.route("/gubreleme", methods=["GET", "POST"])
@login_required
def gubreleme():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            tarih = request.form['tarih']
            sera_adi = request.form['sera_adi'].strip()
            gubre_adi = request.form['gubre_adi'].strip()
            gubre_miktari = float(request.form['gubre_miktari'])
            uygulama_sekli = request.form.get('uygulama_sekli', '').strip()
            personel = request.form['personel'].strip()
            notlar = request.form.get('notlar', '').strip()
            
            if not sera_adi or not gubre_adi or not personel or not tarih or gubre_miktari <= 0:
                flash('Sera adı, gübre adı, personel, tarih ve miktar zorunludur!', 'error')
            else:
                # Automatic stock deduction
                deductions = [('Gübre', gubre_miktari, f'Gübreleme - {tarih} - {sera_adi}')]
                stock_success, stock_results = batch_deduct_stock(conn, deductions)
                
                if stock_success:
                    audit_fields = add_audit_fields_for_create()
                    conn.execute("""
                        INSERT INTO gubreleme (tarih, sera_adi, gubre_adi, gubre_miktari, uygulama_sekli, 
                                             personel, notlar, created_by, modified_by, modified_at) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (tarih, sera_adi, gubre_adi, gubre_miktari, uygulama_sekli, personel, 
                          notlar, audit_fields['created_by'], audit_fields['modified_by'], 
                          audit_fields['modified_at']))
                    conn.commit()
                    
                    success_msg = f'Gübreleme kaydı başarıyla eklendi! {stock_results[0][1]}'
                    flash(success_msg, 'success')
                    return redirect(url_for('gubreleme'))
                else:
                    conn.rollback()
                    flash(f'Stok yetersiz: {stock_results[0][1]}', 'error')
                
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Hata oluştu: {str(e)}', 'error')
            conn.rollback()
        finally:
            conn.close()
    
    # GET request - show fertilization records
    conn = get_db_connection()
    
    # Get fertilization records
    gubrelemeler = conn.execute("""
        SELECT * FROM gubreleme 
        ORDER BY tarih DESC, olusturma_tarihi DESC
    """).fetchall()
    
    # Get monthly statistics
    bu_ay = datetime.now().strftime('%Y-%m')
    bu_ay_gubre = conn.execute("""
        SELECT COALESCE(SUM(gubre_miktari), 0) as total 
        FROM gubreleme 
        WHERE tarih LIKE ?
    """, (f"{bu_ay}%",)).fetchone()['total']
    
    # Get fertilizer type statistics
    gubre_turu_istatistik = conn.execute("""
        SELECT gubre_adi, COUNT(*) as kullanim_sayisi, COALESCE(SUM(gubre_miktari), 0) as toplam_miktar
        FROM gubreleme 
        WHERE tarih LIKE ?
        GROUP BY gubre_adi
        ORDER BY toplam_miktar DESC
    """, (f"{bu_ay}%",)).fetchall()
    
    # Get available fertilizers from stock
    available_fertilizers = conn.execute("""
        SELECT DISTINCT malzeme_adi, miktar, birim
        FROM stok 
        WHERE kategori = 'Gübre' AND miktar > 0
        ORDER BY malzeme_adi
    """).fetchall()
    
    conn.close()
    
    return render_template('gubreleme.html', 
                         gubrelemeler=gubrelemeler,
                         bu_ay_gubre=bu_ay_gubre,
                         gubre_turu_istatistik=gubre_turu_istatistik,
                         available_fertilizers=available_fertilizers)

@app.route("/audit-trail")
@login_required
def audit_trail():
    conn = get_db_connection()
    
    # Recent changes from all tables
    audit_data = []
    
    # Get recent uretim changes
    uretim_changes = conn.execute("""
        SELECT 'Üretim' as tablo, sera_adi || ' - ' || urun_adi as kayit, 
               created_by, modified_by, modified_at, id
        FROM uretim 
        WHERE modified_at IS NOT NULL 
        ORDER BY modified_at DESC LIMIT 10
    """).fetchall()
    audit_data.extend(uretim_changes)
    
    # Get recent personel changes
    personel_changes = conn.execute("""
        SELECT 'Personel' as tablo, personel_adi as kayit,
               created_by, modified_by, modified_at, id
        FROM personel 
        WHERE modified_at IS NOT NULL 
        ORDER BY modified_at DESC LIMIT 10
    """).fetchall()
    audit_data.extend(personel_changes)
    
    # Get recent stok changes
    stok_changes = conn.execute("""
        SELECT 'Stok' as tablo, malzeme_adi as kayit,
               created_by, modified_by, modified_at, id
        FROM stok 
        WHERE modified_at IS NOT NULL 
        ORDER BY modified_at DESC LIMIT 10
    """).fetchall()
    audit_data.extend(stok_changes)
    
    # Get recent devam changes
    devam_changes = conn.execute("""
        SELECT 'Devam' as tablo, p.personel_adi || ' (' || d.tarih || ')' as kayit,
               d.created_by, d.modified_by, d.modified_at, d.id
        FROM devam d
        LEFT JOIN personel p ON d.personel_id = p.id
        WHERE d.modified_at IS NOT NULL 
        ORDER BY d.modified_at DESC LIMIT 10
    """).fetchall()
    audit_data.extend(devam_changes)
    
    # Get recent hasat changes
    hasat_changes = conn.execute("""
        SELECT 'Hasat' as tablo, parsil_alan || ' (' || hasat_tarihi || ')' as kayit,
               created_by, modified_by, modified_at, id
        FROM hasat 
        WHERE modified_at IS NOT NULL 
        ORDER BY modified_at DESC LIMIT 10
    """).fetchall()
    audit_data.extend(hasat_changes)
    
    # Sort all changes by modification time
    audit_data.sort(key=lambda x: x['modified_at'] or '', reverse=True)
    audit_data = audit_data[:50]  # Limit to 50 most recent changes
    
    # Statistics
    stats = {
        'total_changes_today': len([a for a in audit_data if a['modified_at'] and a['modified_at'].startswith(datetime.now().strftime('%Y-%m-%d'))]),
        'active_users_today': len(set([a['modified_by'] for a in audit_data if a['modified_by'] and a['modified_at'] and a['modified_at'].startswith(datetime.now().strftime('%Y-%m-%d'))])),
        'total_records': len(audit_data)
    }
    
    conn.close()
    
    return render_template('audit_trail.html', audit_data=audit_data, stats=stats)

@app.route("/rapor")
@login_required
def rapor():
    conn = get_db_connection()
    
    # Aylık üretim raporu
    uretim_raporu = conn.execute("""
        SELECT strftime('%Y-%m', ekim_tarihi) as ay,
               COUNT(*) as toplam_ekim,
               SUM(CASE WHEN durum = 'Hasat Edildi' THEN 1 ELSE 0 END) as hasat_edilen,
               SUM(CASE WHEN gercek_verim IS NOT NULL THEN gercek_verim ELSE 0 END) as toplam_verim
        FROM uretim 
        WHERE ekim_tarihi >= date('now', '-12 months')
        GROUP BY strftime('%Y-%m', ekim_tarihi)
        ORDER BY ay DESC
    """).fetchall()
    
    # Stok durumu
    stok_raporu = conn.execute("""
        SELECT kategori, 
               COUNT(*) as cesit_sayisi,
               SUM(miktar * maliyet) as toplam_deger
        FROM stok 
        WHERE kategori IS NOT NULL AND kategori != ''
        GROUP BY kategori
        ORDER BY toplam_deger DESC
    """).fetchall()
    
    # Personel maliyeti trendi
    personel_raporu = conn.execute("""
        SELECT strftime('%Y-%m', 'now', '-' || (t.value-1) || ' months') as ay,
               (SELECT COUNT(*) FROM personel WHERE aktif = 1) as personel_sayisi,
               (SELECT SUM(aylik_maas) FROM personel WHERE aktif = 1) as toplam_maliyet
        FROM (SELECT 1 as value UNION SELECT 2 UNION SELECT 3 UNION SELECT 4 UNION SELECT 5 UNION SELECT 6 
              UNION SELECT 7 UNION SELECT 8 UNION SELECT 9 UNION SELECT 10 UNION SELECT 11 UNION SELECT 12) t
        ORDER BY ay DESC
    """).fetchall()
    
    conn.close()
    
    return render_template('rapor.html',
                         uretim_raporu=uretim_raporu,
                         stok_raporu=stok_raporu,
                         personel_raporu=personel_raporu)

# FINANS VE KAR/ZARAR MODÜLÜ
@app.route("/finans")
@login_required
def finans():
    conn = get_db_connection()
    
    # İstatistikler
    stats = {}
    
    # Bekleyen faturalar
    stats['bekleyen_faturalar'] = conn.execute("""
        SELECT COUNT(*) FROM faturalar WHERE durum = 'Beklemede'
    """).fetchone()[0]
    
    # Bu ay toplam gelir (fiyatlandırılmış faturalar)
    bu_ay = datetime.now().strftime('%Y-%m')
    stats['bu_ay_gelir'] = conn.execute("""
        SELECT COALESCE(SUM(toplam_tutar), 0) FROM faturalar 
        WHERE hasat_tarihi LIKE ? AND birim_fiyat IS NOT NULL
    """, (f"{bu_ay}%",)).fetchone()[0]
    
    # Bu ay toplam gider
    stats['bu_ay_gider'] = conn.execute("""
        SELECT COALESCE(SUM(toplam_tutar), 0) FROM maliyetler 
        WHERE tarih LIKE ?
    """, (f"{bu_ay}%",)).fetchone()[0]
    
    # Bu ay kar/zarar
    stats['bu_ay_kar_zarar'] = stats['bu_ay_gelir'] - stats['bu_ay_gider']
    
    # Son faturalar
    son_faturalar = conn.execute("""
        SELECT f.*, h.parsil_alan, h.hasat_eden
        FROM faturalar f
        LEFT JOIN hasat h ON f.hasat_id = h.id
        ORDER BY f.olusturma_tarihi DESC
        LIMIT 10
    """).fetchall()
    
    # Son maliyetler
    son_maliyetler = conn.execute("""
        SELECT * FROM maliyetler 
        ORDER BY olusturma_tarihi DESC
        LIMIT 10
    """).fetchall()
    
    conn.close()
    
    return render_template('finans.html', 
                         stats=stats,
                         son_faturalar=son_faturalar,
                         son_maliyetler=son_maliyetler)

@app.route("/finans/faturalar", methods=["GET", "POST"])
@login_required
def finans_faturalar():
    conn = get_db_connection()
    
    if request.method == "POST":
        action = request.form.get('action')
        
        if action == 'set_price':
            try:
                fatura_id = int(request.form['fatura_id'])
                birim_fiyat = float(request.form['birim_fiyat'])
                notlar = request.form.get('notlar', '').strip()
                
                # Fatura bilgilerini al
                fatura = conn.execute("""
                    SELECT * FROM faturalar WHERE id = ?
                """, (fatura_id,)).fetchone()
                
                if fatura:
                    toplam_tutar = fatura['miktar'] * birim_fiyat
                    audit_fields = add_audit_fields_for_update()
                    
                    conn.execute("""
                        UPDATE faturalar 
                        SET birim_fiyat = ?, toplam_tutar = ?, durum = 'Fiyatlandırıldı',
                            fiyat_belirleme_tarihi = ?, notlar = ?, 
                            modified_by = ?, modified_at = ?
                        WHERE id = ?
                    """, (birim_fiyat, toplam_tutar, datetime.now().strftime('%Y-%m-%d'),
                          notlar, audit_fields['modified_by'], audit_fields['modified_at'], fatura_id))
                    
                    conn.commit()
                    flash(f'Fatura fiyatlandırıldı: {birim_fiyat} TL/kg', 'success')
                else:
                    flash('Fatura bulunamadı!', 'error')
                    
            except (ValueError, KeyError):
                flash('Lütfen geçerli bir fiyat girin!', 'error')
            except Exception as e:
                flash(f'Bir hata oluştu: {str(e)}', 'error')
        
        elif action == 'update_status':
            try:
                fatura_id = int(request.form['fatura_id'])
                yeni_durum = request.form['yeni_durum']
                
                audit_fields = add_audit_fields_for_update()
                update_data = [yeni_durum, audit_fields['modified_by'], audit_fields['modified_at'], fatura_id]
                
                if yeni_durum == 'Faturalandı':
                    fatura_tarihi = datetime.now().strftime('%Y-%m-%d')
                    conn.execute("""
                        UPDATE faturalar 
                        SET durum = ?, fatura_tarihi = ?, modified_by = ?, modified_at = ?
                        WHERE id = ?
                    """, [yeni_durum, fatura_tarihi] + update_data[1:])
                elif yeni_durum == 'Tahsil Edildi':
                    tahsil_tarihi = datetime.now().strftime('%Y-%m-%d')
                    conn.execute("""
                        UPDATE faturalar 
                        SET durum = ?, tahsil_tarihi = ?, modified_by = ?, modified_at = ?
                        WHERE id = ?
                    """, [yeni_durum, tahsil_tarihi] + update_data[1:])
                else:
                    conn.execute("""
                        UPDATE faturalar 
                        SET durum = ?, modified_by = ?, modified_at = ?
                        WHERE id = ?
                    """, update_data)
                
                conn.commit()
                flash(f'Fatura durumu güncellendi: {yeni_durum}', 'success')
                
            except (ValueError, KeyError):
                flash('Geçersiz fatura bilgisi!', 'error')
            except Exception as e:
                flash(f'Bir hata oluştu: {str(e)}', 'error')
        
        return redirect(url_for('finans_faturalar'))
    
    # Fatura listesi
    faturalar = conn.execute("""
        SELECT f.*, h.parsil_alan, h.hasat_eden, h.uretim_id,
               u.sera_adi, u.urun_adi
        FROM faturalar f
        LEFT JOIN hasat h ON f.hasat_id = h.id
        LEFT JOIN uretim u ON h.uretim_id = u.id
        ORDER BY f.olusturma_tarihi DESC
    """).fetchall()
    
    # İstatistikler
    toplam_bekleyen = conn.execute("""
        SELECT COUNT(*), COALESCE(SUM(miktar), 0) 
        FROM faturalar WHERE durum = 'Beklemede'
    """).fetchone()
    
    toplam_fiyatlandirilmis = conn.execute("""
        SELECT COUNT(*), COALESCE(SUM(toplam_tutar), 0) 
        FROM faturalar WHERE durum IN ('Fiyatlandırıldı', 'Faturalandı', 'Tahsil Edildi')
    """).fetchone()
    
    conn.close()
    
    return render_template('finans_faturalar.html', 
                         faturalar=faturalar,
                         toplam_bekleyen=toplam_bekleyen,
                         toplam_fiyatlandirilmis=toplam_fiyatlandirilmis)

@app.route("/finans/maliyetler", methods=["GET", "POST"])
@login_required
def finans_maliyetler():
    conn = get_db_connection()
    
    if request.method == "POST":
        try:
            maliyet_turu = request.form['maliyet_turu'].strip()
            tarih = request.form['tarih']
            aciklama = request.form['aciklama'].strip()
            miktar = float(request.form.get('miktar', 0) or 0)
            birim = request.form.get('birim', '').strip()
            birim_fiyat = float(request.form['birim_fiyat'])
            toplam_tutar = miktar * birim_fiyat if miktar > 0 else birim_fiyat
            sera_adi = request.form.get('sera_adi', '').strip()
            uretim_id = request.form.get('uretim_id') or None
            tedarikci = request.form.get('tedarikci', '').strip()
            fatura_no = request.form.get('fatura_no', '').strip()
            notlar = request.form.get('notlar', '').strip()
            
            if not maliyet_turu or not aciklama or not tarih:
                flash('Maliyet türü, açıklama ve tarih zorunludur!', 'error')
            else:
                audit_fields = add_audit_fields_for_create()
                conn.execute("""
                    INSERT INTO maliyetler (maliyet_turu, tarih, aciklama, miktar, birim, 
                                          birim_fiyat, toplam_tutar, sera_adi, uretim_id, 
                                          tedarikci, fatura_no, notlar, created_by, modified_by, modified_at) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (maliyet_turu, tarih, aciklama, miktar, birim, birim_fiyat, 
                      toplam_tutar, sera_adi, uretim_id, tedarikci, fatura_no, notlar,
                      audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
                conn.commit()
                flash('Maliyet kaydı başarıyla eklendi!', 'success')
                return redirect(url_for('finans_maliyetler'))
                
        except ValueError:
            flash('Lütfen sayısal değerleri doğru formatta girin!', 'error')
        except Exception as e:
            flash(f'Bir hata oluştu: {str(e)}', 'error')
    
    # Maliyet kayıtları
    maliyetler = conn.execute("""
        SELECT m.*, u.sera_adi as uretim_sera_adi, u.urun_adi
        FROM maliyetler m
        LEFT JOIN uretim u ON m.uretim_id = u.id
        ORDER BY m.tarih DESC, m.olusturma_tarihi DESC
    """).fetchall()
    
    # Aktif üretimler
    aktif_uretimler = conn.execute("""
        SELECT id, sera_adi, urun_adi, ekim_tarihi
        FROM uretim 
        WHERE durum != 'Hasat Edildi' 
        ORDER BY sera_adi, urun_adi
    """).fetchall()
    
    # İstatistikler
    bu_ay = datetime.now().strftime('%Y-%m')
    bu_ay_maliyetler = conn.execute("""
        SELECT maliyet_turu, COUNT(*) as adet, SUM(toplam_tutar) as toplam
        FROM maliyetler 
        WHERE tarih LIKE ?
        GROUP BY maliyet_turu
        ORDER BY toplam DESC
    """, (f"{bu_ay}%",)).fetchall()
    
    toplam_maliyet = conn.execute("""
        SELECT COALESCE(SUM(toplam_tutar), 0) FROM maliyetler 
        WHERE tarih LIKE ?
    """, (f"{bu_ay}%",)).fetchone()[0]
    
    conn.close()
    
    return render_template('finans_maliyetler.html', 
                         maliyetler=maliyetler,
                         aktif_uretimler=aktif_uretimler,
                         bu_ay_maliyetler=bu_ay_maliyetler,
                         toplam_maliyet=toplam_maliyet)

@app.route("/finans/kar-zarar")
@login_required  
def finans_kar_zarar():
    conn = get_db_connection()
    
    # Tarih aralığı parametreleri
    baslangic = request.args.get('baslangic', datetime.now().replace(day=1).strftime('%Y-%m-%d'))
    bitis = request.args.get('bitis', datetime.now().strftime('%Y-%m-%d'))
    sera_filter = request.args.get('sera', '')
    
    # Gelir hesaplama (fiyatlandırılmış faturalar)
    gelir_query = """
        SELECT COALESCE(SUM(f.toplam_tutar), 0) as toplam_gelir,
               COUNT(*) as fatura_sayisi,
               COALESCE(SUM(f.miktar), 0) as toplam_kg
        FROM faturalar f
        LEFT JOIN hasat h ON f.hasat_id = h.id
        WHERE f.hasat_tarihi BETWEEN ? AND ?
        AND f.birim_fiyat IS NOT NULL
    """
    
    gelir_params = [baslangic, bitis]
    if sera_filter:
        gelir_query += " AND h.parsil_alan LIKE ?"
        gelir_params.append(f"%{sera_filter}%")
    
    gelir_data = conn.execute(gelir_query, gelir_params).fetchone()
    
    # Gider hesaplama
    gider_query = """
        SELECT COALESCE(SUM(toplam_tutar), 0) as toplam_gider,
               COUNT(*) as maliyet_sayisi
        FROM maliyetler
        WHERE tarih BETWEEN ? AND ?
    """
    
    gider_params = [baslangic, bitis]
    if sera_filter:
        gider_query += " AND sera_adi LIKE ?"
        gider_params.append(f"%{sera_filter}%")
    
    gider_data = conn.execute(gider_query, gider_params).fetchone()
    
    # Maliyet türlerine göre breakdown
    maliyet_breakdown = conn.execute("""
        SELECT maliyet_turu, COUNT(*) as adet, SUM(toplam_tutar) as toplam
        FROM maliyetler
        WHERE tarih BETWEEN ? AND ?
        {} 
        GROUP BY maliyet_turu
        ORDER BY toplam DESC
    """.format("AND sera_adi LIKE ?" if sera_filter else ""), 
    gider_params).fetchall()
    
    # Firma bazında gelir analizi
    firma_analizi = conn.execute("""
        SELECT f.firma_adi, 
               COUNT(*) as teslimat_sayisi,
               SUM(f.miktar) as toplam_kg,
               AVG(f.birim_fiyat) as ortalama_fiyat,
               SUM(f.toplam_tutar) as toplam_gelir
        FROM faturalar f
        LEFT JOIN hasat h ON f.hasat_id = h.id
        WHERE f.hasat_tarihi BETWEEN ? AND ?
        AND f.birim_fiyat IS NOT NULL
        {}
        GROUP BY f.firma_adi
        ORDER BY toplam_gelir DESC
    """.format("AND h.parsil_alan LIKE ?" if sera_filter else ""), 
    gelir_params).fetchall()
    
    # Kar/zarar hesaplama
    net_kar_zarar = gelir_data['toplam_gelir'] - gider_data['toplam_gider']
    karlillik_orani = (net_kar_zarar / gelir_data['toplam_gelir'] * 100) if gelir_data['toplam_gelir'] > 0 else 0
    kg_basina_gelir = gelir_data['toplam_gelir'] / gelir_data['toplam_kg'] if gelir_data['toplam_kg'] > 0 else 0
    kg_basina_maliyet = gider_data['toplam_gider'] / gelir_data['toplam_kg'] if gelir_data['toplam_kg'] > 0 else 0
    kg_basina_kar = kg_basina_gelir - kg_basina_maliyet
    
    # Sonuçları kar_zarar tablosuna kaydet
    try:
        audit_fields = add_audit_fields_for_create()
        conn.execute("""
            INSERT INTO kar_zarar (hesaplama_tarihi, donem_baslangic, donem_bitis, sera_adi,
                                 toplam_gelir, toplam_gider, net_kar_zarar, karlillik_orani,
                                 hasat_miktari, kg_basina_maliyet, kg_basina_gelir, kg_basina_kar,
                                 created_by, modified_by, modified_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (datetime.now().strftime('%Y-%m-%d'), baslangic, bitis, sera_filter or None,
              gelir_data['toplam_gelir'], gider_data['toplam_gider'], net_kar_zarar, karlillik_orani,
              gelir_data['toplam_kg'], kg_basina_maliyet, kg_basina_gelir, kg_basina_kar,
              audit_fields['created_by'], audit_fields['modified_by'], audit_fields['modified_at']))
        conn.commit()
    except Exception as e:
        # Hesaplama kaydı başarısız olsa da rapor gösterilsin
        print(f"Kar/zarar kayıt hatası: {e}")
    
    # Sera listesi (filtre için)
    sera_listesi = conn.execute("""
        SELECT DISTINCT parsil_alan FROM hasat 
        WHERE parsil_alan IS NOT NULL AND parsil_alan != ''
        ORDER BY parsil_alan
    """).fetchall()
    
    conn.close()
    
    return render_template('finans_kar_zarar.html',
                         gelir_data=gelir_data,
                         gider_data=gider_data,
                         net_kar_zarar=net_kar_zarar,
                         karlillik_orani=karlillik_orani,
                         kg_basina_gelir=kg_basina_gelir,
                         kg_basina_maliyet=kg_basina_maliyet,
                         kg_basina_kar=kg_basina_kar,
                         maliyet_breakdown=maliyet_breakdown,
                         firma_analizi=firma_analizi,
                         sera_listesi=sera_listesi,
                         baslangic=baslangic,
                         bitis=bitis,
                         sera_filter=sera_filter)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
